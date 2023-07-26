# -*- coding: utf-8 -*-
# @Time     : 2022/12/4 13:34
# @Author   : JiMing
# @File     : 基本操作.py
# @SoftWare : PyCharm
import os
import openpyxl

"""
一个Excel电子表格文档称为一个工作簿
一个工作簿保存在一个扩展名为.xlsx的文件中
一个工作簿可以包含多个表
用户当前查看的表（或关闭Excel前最后查看的表）称为活动表
在特定行和列的方格称为单元格、格子

简单来说：
xls是excel2003及以前版本所生成的文件格式
xlsx是excel2007及以后版本所生成的文件格式
（excel 2007之后版本可以打开上述两种格式，但是excel2013只能打开xls格式）
"""
# 1：用openpyxl模块打开Excel文档，查看所有sheet表
"""
openpyxl.load_workbook()函数接受文件名，返回一个workbook数据类型的值。
这个workbook对象代表这个Excel文件，这个有点类似File对象代表一个打开的文本文件。
"""

# 设置存储路径
path = r'G:\PythonStudy\excel_study'

# 修改我们excel保存的工作路径为path
os.chdir(path)

# 返回一个wordbook数据类型的值 返回整个excel工作簿
wordbook = openpyxl.load_workbook('英语四级考试居住信息排查表.xlsx')

# 打印Excel工作簿的所有表
print(wordbook.sheetnames)

# 通过sheet名称获取表格
sheet = wordbook['英语四级考场安排表']
print(sheet)

# 使用workbook.active获取活动表 获取当前Excel选中的表格 活动表
sheet_active = wordbook.active
print(sheet_active)

# 获取表格的尺寸
print(sheet.dimensions)  # 获取英语四级考场安排表的数据大小（尺寸）
print(sheet_active.dimensions)  # 获取吉明表格的数据大小

# 获取单元格中的数据
# 方法1： 指定坐标的方式 sheet['A1']
cell1 = sheet['A1']  # 获取A1单元格的数据
cell2 = sheet['A2']  # 获取A2单元格的数据
print(f"{cell1.value} {cell2.value}")  # 通过value属性获取具体内容值

char_list = [chr(value) for value in range(65, 75)]
# print(char_list)
serial_number = []  # 存放数据单元格的列表
for i in char_list:
    for j in range(1, 11):
        cell = sheet[i + str(j)]
        serial_number.append(cell)
print(serial_number)

cnt = 1
for content in serial_number:
    if cnt % 10 == 0:
        print(f"{content.value}", end='\n')
    else:
        print(f"{content.value}", end=' ')
    cnt = cnt + 1

# 方法2：指定行列的方式 sheet.cell(row=, column=)方式
cell_row_1 = sheet.cell(row=1, column=1)   # 获取的位置为第1行，第1列的数据A1
cell_row_2 = sheet.cell(row=2, column=3)   # 获取的位置为第2行，第3列的数据 C2
print(f"{cell_row_1.value} {cell_row_2.value}")
